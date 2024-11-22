# -*- coding: utf-8 -*-
"""

@author: Viktor B

"""

from PyQt6.QtCore import QThread, pyqtSignal, pyqtSlot, QTimer
import numpy as np
import time
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

from .state import State


class MechanicalTest (QThread):
    '''
    General class for tensile tests that includes general methods for hardware
    '''

    #Signal to update matplotlib
    signal_update_charts = pyqtSignal(list)
    #Signal to start/stop tracking marks when test is run
    signal_start_stop_tracking = pyqtSignal(bool)
    signal_save_image = pyqtSignal(str)
    signal_precond_finished = pyqtSignal()
    signal_test_finished = pyqtSignal()
    
    
    _sample_time = 150  #milli seconds
   
    
    def __init__(self, mot_daq):
        super().__init__()

        self._mot_daq = mot_daq

        self._use_video = False #Flag indicating whether marks are recorded
        self._execute = False #Variable indicating if the test should continue running
        self._state = State.PRECONDITIONING
        
        self._init_variables() 
          
    
    #def __del__(self):
        #self.quit()
        
        
    def _init_variables(self):
        '''
        Function to initialize variables before each measurement

        Returns
        -------
        None.

        '''

        

        # Variables to store measuremetns
        self._load1 = [] #channel1 from load cell
        self._load2 = [] #channel2
        self._disp1 = [] #length
        self._disp2 = [] #length
        self._time_abs = [] #time
        self._time_rel = [] #time
        self._counter = 0 #counter
        
        self._vel_1 = []
        self._vel_2 = []
        
        self._force1 = self._force2 = 0
        self._pos1 = self._pos2 = 0

        if self._use_video:
            self._E11 = []
            self._E12 = []
            self._E22 = []

            self._F11 = []
            self._F12 = []
            self._F21 = []
            self._F22 = []

            self.init_markers()
        
        #record start time
        self._start_cycle_time = time.perf_counter()
    
        
    
    def run(self):
        #Generation of random signal to test the class
        #This method is redefined in subclassess
        while (self._counter < 100):
            self._sendRandSignal()
            QThread.msleep(100)
    
    @pyqtSlot()
    def stop(self):
        """
        Stops movemets, connections and  the execution of the current QThread
        """
        self._execute = False

        if hasattr(self, "_test_timer") and isinstance(self._test_timer, QTimer):
            self._test_timer.stop()

        self.quit()
        self.wait()
    
    def stop_measurement(self):
        """
        Stop the measurement by setting the _execute flag to False and stopping axis1 and axis2.
        """
        self._execute = False
        self.signal_start_stop_tracking.emit(False)
        self._mot_daq.stop_motors()
        
    
    
    @pyqtSlot(list)
    def init_markers(self, array = None):

        print("Markers initialized")
        self._use_video = True

        #datastructures to store tracks of marks
        self._marks_groups = []
        self._point1 = []
        self._point2 = []
        self._point3 = []
        self._point4 = []
        
        self._marks_groups.append(self._point1)
        self._marks_groups.append(self._point2)
        self._marks_groups.append(self._point3)
        self._marks_groups.append(self._point4)

        if array is not None:
            #Temporary coordinates 
            self._temp_p1 = array[0]
            self._temp_p2 = array[1]
            self._temp_p3 = array[2]
            self._temp_p4 = array[3]

        self._p1_0 = self._temp_p1
        self._p2_0 = self._temp_p2
        self._p3_0 = self._temp_p3
        self._p4_0 = self._temp_p4

        self._point1.append(self._temp_p1)
        self._point2.append(self._temp_p2)
        self._point3.append(self._temp_p3)
        self._point4.append(self._temp_p4)
    
    @pyqtSlot(list)
    def update_markers(self, array):

        #print("Markers updated")

        #Temporary coordinates 
        self._temp_p1 = array[0]
        self._temp_p2 = array[1]
        self._temp_p3 = array[2]
        self._temp_p4 = array[3]

    
 
    
    def _update_arrays_emit_data(self):

        """
        Edge and node arrangement for central quad
                     edge 3
                3_______________4
                |               |
                |               |
        edge 2  |               | edge 4
                |               |
                |_______________|
                2    edge 1     1
        """

        #Append common variables for all tests
        
        roundDecimals = 3
        #Add them to current variables
        self._time_abs.append(round(self._current_time, roundDecimals))
        self._time_rel.append(round(self._current_cycle_time, roundDecimals))
        self._load1.append(self._force1) #force channel 1
        self._load2.append(self._force2)
        self._disp1.append(round(-2*self._pos1, roundDecimals)) #Movement of sample is twice as holder
        self._disp2.append(round(-2*self._pos2, roundDecimals))
        self._vel_1.append(self._vel_ax1)
        self._vel_2.append(self._vel_ax2)
        

        if not self._use_video:
            self.signal_update_charts.emit([self._time_abs[-1], self._load1[-1], self._load2[-1], self._disp1[-1], self._disp2[-1]])
        else:
            #calculate strain
            #along horizontal axis - upper and lower groups]
            #the algorithm find contours arranges points along y axis of an image

            #Append points from a temporary buffer that is updated though signal/slot mechanism
            self._point1.append(self._temp_p1)
            self._point2.append(self._temp_p2)
            self._point3.append(self._temp_p3)
            self._point4.append(self._temp_p4)

            
            #Coordinates of initial and final points
            p1 = self._point1[-1]
            p2 = self._point2[-1]
            p3 = self._point3[-1]
            p4 = self._point4[-1]

            

            #decomposed initial coordinates for each node by axis
            X11 = self._p1_0[0]; X21 = self._p1_0[1]
            X12 = self._p2_0[0]; X22 = self._p2_0[1]
            X13 = self._p3_0[0]; X23 = self._p3_0[1]
            X14 = self._p4_0[0]; X24 = self._p4_0[1]

            # component X of displacement on isoparametric coordinates
            dX1_dS1 = 0.25*(X11-X12-X13+X14)
            dX1_dS2 = 0.25*(X11+X12-X13-X14)
            # component Y of displacement on isoparametric coordinates
            dX2_dS1 = 0.25*(X21-X22-X23+X24)
            dX2_dS2 = 0.25*(X21+X22-X23-X24)
            # jacobian of the initial with isoparametric coordinates
            J = dX1_dS1*dX2_dS2 - dX1_dS2*dX2_dS1

            #displacement for each node
            u1 = tuple(a - b for a, b in zip(p1, self._p1_0))
            u2 = tuple(a - b for a, b in zip(p2, self._p2_0))
            u3 = tuple(a - b for a, b in zip(p3, self._p3_0))
            u4 = tuple(a - b for a, b in zip(p4, self._p4_0))

            #decomposed displacement for each node by axis
            u11 = u1[0]; u21 = u1[1]
            u12 = u2[0]; u22 = u2[1]
            u13 = u3[0]; u23 = u3[1]
            u14 = u4[0]; u24 = u4[1]

            # component X of displacement on isoparametric coordinates
            du1_dS1 = 0.25*(u11-u12-u13+u14)
            du1_dS2 = 0.25*(u11+u12-u13-u14)
            # component Y of displacement on isoparametric coordinates
            du2_dS1 = 0.25*(u21-u22-u23+u24)
            du2_dS2 = 0.25*(u21+u22-u23-u24)

            # gradient of displacement u1 on spatial coordinates
            du1_dX1 = (dX2_dS2*du1_dS1 - dX2_dS1*du1_dS2)/J
            du1_dX2 = (-dX1_dS2*du1_dS1 + dX1_dS1*du1_dS2)/J
            # gradient of displacement u2 on spatial coordinates
            du2_dX1 = (dX2_dS2*du2_dS1 - dX2_dS1*du2_dS2)/J
            du2_dX2 = (-dX1_dS2*du2_dS1 + dX1_dS1*du2_dS2)/J

            # deformation gradient, F
            F11 = du1_dX1 + 1.0                # lambda1
            F12 = du1_dX2                      # kappa1
            F21 = du2_dX1                      # kappa2
            F22 = du2_dX2 + 1.0                # lambda2
            F33 = 1.0/(F11*F22 - F12*F21)      # lambda3

            E11 = 0.5*(F11**2 + F21**2 - 1.0)
            E22 = 0.5*(F22**2 + F12**2 - 1.0)
            E12 = 0.5*(F11*F12 + F22*F21)
            

            
            roundDecimals = 4
            
            self._E11.append(round(E11,roundDecimals))
            self._E12.append(round(E12,roundDecimals))
            self._E22.append(round(E22,roundDecimals))

            self._F11.append(round(F11,roundDecimals))
            self._F12.append(round(F12,roundDecimals))
            self._F21.append(round(F21,roundDecimals))
            self._F22.append(round(F22,roundDecimals))

            self.signal_update_charts.emit([self._time_abs[-1], self._load1[-1], self._load2[-1], self._disp1[-1], self._disp2[-1], self._E11[-1], self._E22[-1], self._vel_1[-1],self._vel_2[-1]])

    
    def _finish_test(self):
        
        self._use_video = False
        
        self._init_variables()     
    
    def _moving_average(self, x, w):
        return np.convolve(x, np.ones(w), 'valid') / w
    
    def changeFolder(self, folder, sam_name):
        self._workfolder = folder
        self._sam_name = sam_name
        

    def _writeDataToFile(self, file_name, cycl_num, fl_txt):
        
        #Zip data in one variable for export
        if self._use_video:
            X11	= [x for x,_ in self._point1]
            X12	= [x for x,_ in self._point2]
            X13	= [x for x,_ in self._point3]
            X14	= [x for x,_ in self._point4]
            X21	= [y for _,y in self._point1]
            X22	= [y for _,y in self._point2]
            X23	= [y for _,y in self._point3]
            X24 = [y for _,y in self._point4]

            header = ["Time_rel", "Load_1","Load_2", "F11","F12","F21","F22", "E11","E12","E22", "X11","X12","X13","X14","X21","X22","X23","X24", "Disp_1", "Disp_2", "Time_abs"]
            combined_lists = list(zip(self._time_rel, self._load1, self._load2, 
                                 self._F11, self._F12, self._F21, self._F22,
                                 self._E11, self._E12, self._E22, X11,X12,X13,X14,X21,X22,X23,X24, 
                                 self._disp1, self._disp2, self._time_abs))
        
        else: 
            header = ["Time_rel", "Load_1", "Load_2", "Disp_1", "Disp_2", "Time_abs"]
            combined_lists = list(zip( self._time_rel, self._load1, self._load2, self._disp1, self._disp2, self._time_abs))

        folder = os.path.join(self._workfolder, self._sam_name)
        os.makedirs(folder, exist_ok=True)

        #Write in excel file
        file_path = os.path.join(folder, f'{file_name}.xlsx')

        # Check if the file exists
        if os.path.exists(file_path):
            # Load the existing workbook
            wb = load_workbook(file_path)
        else:
            # Create a new workbook if it doesn't exist
            wb = Workbook()
            # Remove the default sheet created by openpyxl
            default_sheet = wb.active
            wb.remove(default_sheet)

        # Add a new sheet with a specific name
        ws = wb.create_sheet(title=f'Cycl_{cycl_num}')
        # Set the header row
        ws.append(header)
        # Write the data rows
        for row in combined_lists:
            ws.append(row)

        # Save the workbook to the specified file
        wb.save(file_path)

        #If we need to save the file also in txt format
        if fl_txt:
            txt_file_path = os.path.join(folder, f'{file_name}.txt')
            try:
                with open(txt_file_path, 'w', encoding='utf-8') as f:
                    # Write header
                    header_line = '\t'.join(header)
                    f.write(header_line + '\n')
                    print(f"Written header to text file: {header_line}")

                    # Write data rows
                    for idx, row in enumerate(combined_lists):
                        # Ensure each element is a string and join with tabs
                        row_str = '\t'.join(map(str, row))
                        f.write(row_str + '\n')
                        if idx <= 5:  # Print first 5 rows for debugging
                            print(f"Written row {idx} to text file: {row_str}")

                print(f"Data successfully written to text file: {txt_file_path}")
            except Exception as e:
                print(f"Error writing to text file: {e}")


    def _write_points_reference(self, folder):
        if self._use_video:
            # Format the date and time with tabulation
            now = datetime.now()
            formatted_time = now.strftime("%m/%d/%Y\t%I:%M %p")

            # Validate that temporary points have at least two coordinates
            if not all(len(p) >= 2 for p in [self._temp_p1, self._temp_p2, self._temp_p3, self._temp_p4]):
                print("_write_points_reference: One or more temporary points do not contain enough coordinates.")
                return

            # Extract X and Y coordinates
            x_coords = [int(p[0]) for p in [self._temp_p1, self._temp_p2, self._temp_p3, self._temp_p4]]
            y_coords = [int(p[1]) for p in [self._temp_p1, self._temp_p2, self._temp_p3, self._temp_p4]]

            combined_lists = [formatted_time] + x_coords + y_coords


            txt_file_path = os.path.join(folder, f'REFERENCE_LOG.txt')
            try:
                with open(txt_file_path, 'a', encoding='utf-8') as f:

                    row_str = '\t'.join(map(str, combined_lists))
                    f.write(row_str + '\n')

                print(f"Updating REFERENCE_LOG.txt\n{row_str}")
            except Exception as e:
                print(f"Error writing to text file: {e}")

            

    def _send_rand_signal(self):
       
        print("Sending random")
        self._counter +=1
        self._load1 = np.random.rand(1)
        self._load2 = np.random.rand(1) 
        self._disp1 = np.random.rand(1) 
        self._disp2 = np.random.rand(1) 
        self._E11 = np.random.rand(1) 
        self._E22 = np.random.rand(1) 
        self._vel_1 = np.random.rand(1) 
        self._vel_2 = np.random.rand(1) 
        self.signal_update_charts.emit([self._counter, self._load1[-1], self._load2[-1], self._disp1[-1], self._disp2[-1], self._E11[-1], self._E22[-1], self._vel_1[-1],self._vel_2[-1]])



        
            
            
    
