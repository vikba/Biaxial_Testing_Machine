# -*- coding: utf-8 -*-
"""

@author: Viktor B

"""

from PyQt6.QtCore import QThread, pyqtSignal, pyqtSlot,  Qt, QTimer, QMetaObject
import numpy as np
import csv
import time
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

from .state import State
from .direction import Direction


class MechanicalTest (QThread):
    '''
    General class for tensile tests that includes general methods for hardware
    '''

    #Signal to update matplotlib
    signal_update_charts = pyqtSignal(list)
    #Signal to start/stop tracking marks when test is run
    signal_start_stop_tracking = pyqtSignal(bool)
    signal_save_image = pyqtSignal(str)
    signal_precond_finished = pyqtSignal()
    signal_test_finished = pyqtSignal()
    
    
    _sample_time = 150  #milli seconds
   
    
    def __init__(self, mot_daq):
        super().__init__()

        self._mot_daq = mot_daq

        self._use_video = False #Flag indicating whether marks are recorded
        self._execute = True #Variable indicating if the test should continue running
        self._state = State.PRECONDITIONING
        
        self._init_variables() 
          
    
    #def __del__(self):
        #self.quit()
        
        
    def _init_variables(self):
        '''
        Function to initialize variables before each measurement

        Returns
        -------
        None.

        '''

        

        # Variables to store measuremetns
        self._load1 = [] #channel1 from load cell
        self._load2 = [] #channel2
        self._disp1 = [] #length
        self._disp2 = [] #length
        self._time_abs = [] #time
        self._time_rel = [] #time
        self._counter = 0 #counter
        
        self._vel_1 = []
        self._vel_2 = []
        
        self._force1 = self._force2 = 0
        self._pos1 = self._pos2 = 0

        if self._use_video:
            self._E11 = []
            self._E12 = []
            self._E22 = []

            self._F11 = []
            self._F12 = []
            self._F21 = []
            self._F22 = []

            self.init_markers()
        
        #record start time
        self._start_cycle_time = time.perf_counter()
    
        
    
    def run(self):
        #Generation of random signal to test the class
        #This method is redefined in subclassess
        while (self._counter < 100):
            self._sendRandSignal()
            QThread.msleep(100)
    
    @pyqtSlot()
    def stop(self):
        """
        Stops movemets, connections and  the execution of the current QThread
        """
        self._execute = False

        if hasattr(self, "_test_timer") and isinstance(self._test_timer, QTimer):
            self._test_timer.stop()

        self.quit()
        self.wait()
    
    def stop_measurement(self):
        """
        Stop the measurement by setting the _execute flag to False and stopping axis1 and axis2.
        """
        self._execute = False
        self.signal_start_stop_tracking.emit(False)
        self._mot_daq.stop_motors()
        
    
    
    @pyqtSlot(list)
    def init_markers(self, array = None):

        print("Markers initialized")
        self._use_video = True

        #datastructures to store tracks of marks
        self._marks_groups = []
        self._point1 = []
        self._point2 = []
        self._point3 = []
        self._point4 = []
        
        self._marks_groups.append(self._point1)
        self._marks_groups.append(self._point2)
        self._marks_groups.append(self._point3)
        self._marks_groups.append(self._point4)

        if array is not None:
            #Temporary coordinates 
            self._temp_p1 = array[0]
            self._temp_p2 = array[1]
            self._temp_p3 = array[2]
            self._temp_p4 = array[3]

        self._p1_0 = self._temp_p1
        self._p2_0 = self._temp_p2
        self._p3_0 = self._temp_p3
        self._p4_0 = self._temp_p4

        self._point1.append(self._temp_p1)
        self._point2.append(self._temp_p2)
        self._point3.append(self._temp_p3)
        self._point4.append(self._temp_p4)
    
    @pyqtSlot(list)
    def update_markers(self, array):

        #print("Markers updated")

        #Temporary coordinates 
        self._temp_p1 = array[0]
        self._temp_p2 = array[1]
        self._temp_p3 = array[2]
        self._temp_p4 = array[3]

    
 
    
    def _update_arrays_emit_data(self):

        """
        Edge and node arrangement for central quad
                     edge 3
                3_______________4
                |               |
                |               |
        edge 2  |               | edge 4
                |               |
                |_______________|
                2    edge 1     1
        """

        #Append common variables for all tests
        
        roundDecimals = 3
        #Add them to current variables
        self._time_abs.append(round(self._current_time, roundDecimals))
        self._time_rel.append(round(self._current_cycle_time, roundDecimals))
        self._load1.append(self._force1) #force channel 1
        self._load2.append(self._force2)
        self._disp1.append(round(-2*self._pos1, roundDecimals)) #Movement of sample is twice as holder
        self._disp2.append(round(-2*self._pos2, roundDecimals))
        self._vel_1.append(self._vel_ax1)
        self._vel_2.append(self._vel_ax2)
        

        if not self._use_video:
            self.signal_update_charts.emit([self._time_abs[-1], self._load1[-1], self._load2[-1], self._disp1[-1], self._disp2[-1]])
        else:
            #calculate strain
            #along horizontal axis - upper and lower groups]
            #the algorithm find contours arranges points along y axis of an image

            #Append points from a temporary buffer that is updated though signal/slot mechanism
            self._point1.append(self._temp_p1)
            self._point2.append(self._temp_p2)
            self._point3.append(self._temp_p3)
            self._point4.append(self._temp_p4)

            
            #Coordinates of initial and final points
            p1 = self._point1[-1]
            p2 = self._point2[-1]
            p3 = self._point3[-1]
            p4 = self._point4[-1]

            

            #decomposed initial coordinates for each node by axis
            X11 = self._p1_0[0]; X21 = self._p1_0[1]
            X12 = self._p2_0[0]; X22 = self._p2_0[1]
            X13 = self._p3_0[0]; X23 = self._p3_0[1]
            X14 = self._p4_0[0]; X24 = self._p4_0[1]

            # component X of displacement on isoparametric coordinates
            dX1_dS1 = 0.25*(X11-X12-X13+X14)
            dX1_dS2 = 0.25*(X11+X12-X13-X14)
            # component Y of displacement on isoparametric coordinates
            dX2_dS1 = 0.25*(X21-X22-X23+X24)
            dX2_dS2 = 0.25*(X21+X22-X23-X24)
            # jacobian of the initial with isoparametric coordinates
            J = dX1_dS1*dX2_dS2 - dX1_dS2*dX2_dS1

            #displacement for each node
            u1 = tuple(a - b for a, b in zip(p1, self._p1_0))
            u2 = tuple(a - b for a, b in zip(p2, self._p2_0))
            u3 = tuple(a - b for a, b in zip(p3, self._p3_0))
            u4 = tuple(a - b for a, b in zip(p4, self._p4_0))

            #decomposed displacement for each node by axis
            u11 = u1[0]; u21 = u1[1]
            u12 = u2[0]; u22 = u2[1]
            u13 = u3[0]; u23 = u3[1]
            u14 = u4[0]; u24 = u4[1]

            # component X of displacement on isoparametric coordinates
            du1_dS1 = 0.25*(u11-u12-u13+u14)
            du1_dS2 = 0.25*(u11+u12-u13-u14)
            # component Y of displacement on isoparametric coordinates
            du2_dS1 = 0.25*(u21-u22-u23+u24)
            du2_dS2 = 0.25*(u21+u22-u23-u24)

            # gradient of displacement u1 on spatial coordinates
            du1_dX1 = (dX2_dS2*du1_dS1 - dX2_dS1*du1_dS2)/J
            du1_dX2 = (-dX1_dS2*du1_dS1 + dX1_dS1*du1_dS2)/J
            # gradient of displacement u2 on spatial coordinates
            du2_dX1 = (dX2_dS2*du2_dS1 - dX2_dS1*du2_dS2)/J
            du2_dX2 = (-dX1_dS2*du2_dS1 + dX1_dS1*du2_dS2)/J

            # deformation gradient, F
            F11 = du1_dX1 + 1.0                # lambda1
            F12 = du1_dX2                      # kappa1
            F21 = du2_dX1                      # kappa2
            F22 = du2_dX2 + 1.0                # lambda2
            F33 = 1.0/(F11*F22 - F12*F21)      # lambda3

            E11 = 0.5*(F11**2 + F21**2 - 1.0)
            E22 = 0.5*(F22**2 + F12**2 - 1.0)
            E12 = 0.5*(F11*F12 + F22*F21)
            

            

            '''
            #Old formula for strain calculation
            L1_0 = (math.dist(p1_0, p2_0) + math.dist(p3_0,p4_0)) / 2
            L1_last = (math.dist(p1, p2) + math.dist(p3, p4)) / 2

            if (p1_0[1]-p3_0[1] < p1_0[1]-p4_0[1]):
                L2_0 = (math.dist(p1_0, p3_0) + math.dist(p2_0,p4_0)) / 2
                L2_last = (math.dist(p1, p3) + math.dist(p2, p4)) / 2
            else:
                L2_0 = (math.dist(p1_0, p4_0) + math.dist(p2_0, p3_0)) / 2
                L2_last = (math.dist(p1, p4) + math.dist(p2, p3)) / 2

            # Calculate strains
            lambda1 = 1 + (L1_last - L1_0) / L1_0
            lambda2 = 1 + (L2_last - L2_0) / L2_0

            E11 = (lambda1*lambda1 - 1)/2
            E22 = (lambda2*lambda2 - 1)/2'''

            
            roundDecimals = 4
            
            self._E11.append(round(E11,roundDecimals))
            self._E12.append(round(E12,roundDecimals))
            self._E22.append(round(E22,roundDecimals))

            self._F11.append(round(F11,roundDecimals))
            self._F12.append(round(F12,roundDecimals))
            self._F21.append(round(F21,roundDecimals))
            self._F22.append(round(F22,roundDecimals))

            self.signal_update_charts.emit([self._time_abs[-1], self._load1[-1], self._load2[-1], self._disp1[-1], self._disp2[-1], self._E11[-1], self._E22[-1], self._vel_1[-1],self._vel_2[-1]])


    def _finish_test(self):
        
        self._use_video = False
        
        self._init_variables()     
    
    def _moving_average(self, x, w):
        return np.convolve(x, np.ones(w), 'valid') / w
    
    def changeFolder(self, folder, sam_name):
        self._workfolder = folder
        self._sam_name = sam_name
        

    def _writeDataToFile(self, file_name, cycl_num):
        # Combine the lists

        if self._use_video:
            X11	= [x for x,_ in self._point1]
            X12	= [x for x,_ in self._point2]
            X13	= [x for x,_ in self._point3]
            X14	= [x for x,_ in self._point4]
            X21	= [y for _,y in self._point1]
            X22	= [y for _,y in self._point2]
            X23	= [y for _,y in self._point3]
            X24 = [y for _,y in self._point4]

            header = ["Time_rel", "Load_1","Load_2", "F11","F12","F21","F22", "E11","E12","E22", "X11","X12","X13","X14","X21","X22","X23","X24", "Disp_1", "Disp_2", "Time_abs"]
            combined_lists = zip(self._time_rel, self._load1, self._load2, 
                                 self._F11, self._F12, self._F21, self._F22,
                                 self._E11, self._E12, self._E22, X11,X12,X13,X14,X21,X22,X23,X24, 
                                 self._disp1, self._disp2, self._time_abs)
        
        else: 
            header = ["Time_rel", "Load_1", "Load_2", "Disp_1", "Disp_2", "Time_abs"]
            combined_lists = zip( self._time_rel, self._load1, self._load2, self._disp1, self._disp2, self._time_abs)

        folder = os.path.join(self._workfolder, self._sam_name)
        os.makedirs(folder, exist_ok=True)

        file_path = os.path.join(folder, f'{file_name}.xlsx')

        # Check if the file exists
        if os.path.exists(file_path):
            # Load the existing workbook
            wb = load_workbook(file_path)
        else:
            # Create a new workbook if it doesn't exist
            wb = Workbook()
            # Remove the default sheet created by openpyxl
            default_sheet = wb.active
            wb.remove(default_sheet)

        # Add a new sheet with a specific name
        ws = wb.create_sheet(title=f'Cycl_{cycl_num}')
        # Set the header row
        ws.append(header)
        # Write the data rows
        for row in combined_lists:
            ws.append(row)

        # Save the workbook to the specified file
        wb.save(file_path)

        '''# Write to file positions of markers
        if self._use_video:
            file_path = os.path.join(self._workfolder, file_name + "_markers.xlsx")
            if os.path.exists(file_path):
                wb_mark = load_workbook(file_path)
            else:
                wb_mark = Workbook()
                default_sheet = wb_mark.active
                wb_mark.remove(default_sheet)

            combined_lists = zip(self._time_rel, self._point1[-1][0], self._point1[-1][1], self._point2[-1][0], self._point2[-1][1], self._point3[-1][0], self._point3[-1][1], self._point4[-1][0], self._point4[-1][1])
            ws_mark = wb_mark.create_sheet(title=f'Cycl_{cycl_num}')
            ws_mark.append(["Time","Marker_1_X", "Marker_1_Y", "Marker_2_X", "Marker_2_Y", "Marker_3_X", "Marker_3_Y", "Marker_4_X", "Marker_4_Y"])

            for row in combined_lists:
                ws_mark.append(row)

            wb_mark.save(file_path)'''

        
        '''# Write to CSV file
        with open(os.path.join(self._workfolder, f'{str}_{formatted_datetime}_({num}).csv'), 'w', newline='') as file:
            writer = csv.writer(file, delimiter=',')
            writer.writerow(["Time", "Load_1", "Load_2", "Disp_1", "Disp_2", "E11", "E22"])  # Header row, if needed
            for row in combined_lists:
                writer.writerow(row)
                
        # If videoextensometer is on and more than 10 marker positions were recorded
        # Write to CSV positions of markers
        if self._use_video:
            combined_lists = zip(self._time, self._point1, self._point2, self._point3, self._point4)
            with open(os.path.join(self._workfolder, f'{str}_{formatted_datetime}_({num})_markers.csv'), 'w', newline='') as file:
                writer = csv.writer(file, delimiter=',')
                writer.writerow(["Time","Marker_1", "Marker_2", "Marker_3", "Marker_4"])  # Header row, if needed
                for row in combined_lists:
                    writer.writerow(row)'''

            

    def _send_rand_signal(self):
       
        print("Sending random")
        self._counter +=1
        self._load1 = np.random.rand(1)
        self._load2 = np.random.rand(1) 
        self._disp1 = np.random.rand(1) 
        self._disp2 = np.random.rand(1) 
        self._E11 = np.random.rand(1) 
        self._E22 = np.random.rand(1) 
        self._vel_1 = np.random.rand(1) 
        self._vel_2 = np.random.rand(1) 
        self.signal_update_charts.emit([self._counter, self._load1[-1], self._load2[-1], self._disp1[-1], self._disp2[-1], self._E11[-1], self._E22[-1], self._vel_1[-1],self._vel_2[-1]])


class DisplacementControlTest(MechanicalTest):
    
     
    def __init__(self, mot_daq, folder, sam_name, vel1, vel2, len1, len2, num_cycles):
        super().__init__(mot_daq)

        self._mot_daq = mot_daq
        
        # Set speed
        self._vel_ax1 = vel1
        self._vel_ax2 = vel2

        self._max_disp1 = len1
        self._max_disp2 = len2
        
        self._workfolder = folder
        self._sam_name = sam_name

        self._num_cycles = num_cycles


        #this timers should be created in class, but not in its parent class
        
    def update_parameters(self, sam_name, vel1, vel2, len1, len2, num_cycles):

        self._sam_name = sam_name
        
        # Set speed
        self._vel_ax1 = vel1
        self._vel_ax2 = vel2

        self._max_disp1 = len1
        self._max_disp2 = len2

        self._num_cycles = num_cycles
        
        
    def run(self):
        """
        This function is implemented when the QTHread starts. Implements displacement control test.

        This function stops the liveforce timer, initializes the variables, starts tracking, 
        moves the motors with specified velocities, and performs a measurement cycle.

        """
        
        self._init_variables()

        if (self._use_video):
            print(f"Start of Displacement control test with video")
            self.signal_start_stop_tracking.emit(True)
            
            #Get current date for filename
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M")
            img_addr = self._workfolder + '\\Test_'+self._sam_name+'_'+formatted_datetime+'_first_frame.jpg'
            self.signal_save_image.emit(img_addr)
        
        #Here should be a condition to start the test
        if True:

            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M")
            self._file_name = 'DTest_'+ self._sam_name + '_' + formatted_datetime
            
            #Set number of cycles and direction
            self._half_cycle = 2*self._num_cycles #double for each half cycle
            self._direction = Direction.STRETCH

            self._start_time = time.perf_counter()
            self._start_cycle_time = time.perf_counter()
            self._current_time = 0

            #Set final positions
            self._pos1, self._pos2 = self._mot_daq.get_positions()

            self._start_pos1 = self._pos1
            self._start_pos2 = self._pos2
            self._fin_pos1 = self._start_pos1 - self._max_disp1/2 #when sample is stretched, the position is decreased
            self._fin_pos2 = self._start_pos2 - self._max_disp2/2

            #Start movement to final position
            self._mot_daq.move_position_ax1(self._fin_pos1, self._vel_ax1)
            self._mot_daq.move_position_ax2(self._fin_pos2, self._vel_ax2)

            #Start timer to periodically check length and control the test
            
            self._test_timer = QTimer()
            self._test_timer.timeout.connect(self._one_cycle)
            #self._test_timer.moveToThread(self)
            self._test_timer.start(self._sample_time)
            print("Timer started")
            self.exec()
            
        
        

    def _one_cycle(self):

        #Condition to finish the test: positive number of steps left
        if self._current_time < 1000 and self._execute and self._half_cycle > 0:

            #Update variables
            self._pos1, self._pos2 = self._mot_daq.get_positions()
            self._force1,self._force2 = self._mot_daq.get_forces()
            self._current_time = time.perf_counter() - self._start_time
            self._current_cycle_time = time.perf_counter() - self._start_cycle_time

            #If Stretch or Relax continue - send data 
            if Direction.STRETCH == self._direction and (self._pos1 >= self._fin_pos1+0.08 or self._pos2 >= self._fin_pos2+0.08) or \
                Direction.COMPRESS == self._direction and (self._pos1 < self._fin_pos1-0.08 or self._pos2 < self._fin_pos2-0.08):

                self._update_arrays_emit_data()

                #print(f"positions current: {self._pos1} , final: {self._fin_pos1}")
                #print(f"positions current: {self._pos2} , final: {self._fin_pos2}")
                
            #If Half cycle finished - change direction
            else:
                print("DisplacementControlTest: Half cycle finished")
                self._half_cycle -= 1
                if Direction.STRETCH == self._direction:
                    self._direction = Direction.COMPRESS
                else:
                    self._direction = Direction.STRETCH

                #Set final posistion based on stretch or relax cycle
                #Stretch cycle
                if Direction.STRETCH == self._direction:
                    self._fin_pos1 = self._start_pos1 - self._max_disp1/2 
                    self._fin_pos2 = self._start_pos2 - self._max_disp2/2

                    self._writeDataToFile(self._file_name, self._half_cycle/2 +1)
                    self._start_cycle_time = time.perf_counter()

                #Relax cycle
                elif Direction.COMPRESS == self._direction:
                    self._fin_pos1 = self._start_pos1
                    self._fin_pos2 = self._start_pos2

                if (self._use_video):
                    #Get current date for filename
                    current_datetime = datetime.now()
                    formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M")
                    img_addr = self._workfolder + '\\Test_'+self._sam_name+'_'+formatted_datetime+'_last_frame.jpg'
                    self.signal_save_image.emit(img_addr)
                
                if self._half_cycle > 0:
                    print(f"positions current: {self._pos1} , final: {self._fin_pos1}")
                    print(f"positions current: {self._pos2} , final: {self._fin_pos2}")
                    
                    #Start movement to final position
                    self._mot_daq.move_position_ax1(self._fin_pos1, self._vel_ax1)
                    self._mot_daq.move_position_ax2(self._fin_pos2, self._vel_ax2)
        
        #Test finished
        else:
            # Stop motors after measurement cycle is finished
            #self._test_timer.stop()
            QMetaObject.invokeMethod(self._test_timer, "stop", Qt.ConnectionType.QueuedConnection)
            print("DisplacementControlTest: Test finished")
            self.stop_measurement()
            self._finish_test()
            
        
        
            
            

class LoadControlTest(MechanicalTest):
    
    '''
    num_cycles:
    0 - one directional (half cycle)
    1 - one cycle
    etc
    '''

    def __init__(self, mot_daq, folder, sam_name, max_force1, max_force2, disp1, disp2, test_duration, num_cycles_precond, num_cycles_test):
        super().__init__(mot_daq)
        self._mot_daq = mot_daq
        
        self._max_force1 = max_force1
        self._max_force2 = max_force2
        self._test_duration = test_duration
        
        self._workfolder = folder
        self._sam_name = sam_name

        self._num_cycles_precond = num_cycles_precond
        self._num_cycles_test = num_cycles_test

        self._disp_guess1 = disp1
        self._disp_guess2 = disp2
    
    def update_parameters (self, work_folder, sam_name, end_force1, end_force2, disp_guess1, disp_guess2, test_duration, num_cycles_precond, num_cycles_test):
        self._workfolder = work_folder
        self._sam_name = sam_name
        self._end_force1 = end_force1
        self._end_force2 = end_force2
        self._disp_guess1 = disp_guess1
        self._disp_guess2 = disp_guess2
        self._test_duration = test_duration
        self._num_cycles_precond = num_cycles_precond
        self._num_cycles_test = num_cycles_test
    
    def increase_desired_force(self, start_force, end_force, duration, current_time):
        """
        Calculates desired force over the specified duration for specific current_time.
        Important function to control PID behaviour
    
        :param start_force: Initial force in Newtons
        :param end_force: Final force in Newtons
        :param duration: Duration of the test in seconds
        :param self._current_time: Current time in seconds since the start of the test
        :return: Desired force at the current time
        """
        if current_time >= duration:
            return end_force
        return start_force + (end_force - start_force) * (current_time / duration)
    



    def run(self):
        """
        This function is implemented when the QTHread starts. Implements load control test.
        """
        

        #start performing test
        #this function is required to perform test with or without camera

        self._init_variables()

        print(f"use video: {self._use_video}")

        #Get current date for filename
        current_datetime = datetime.now()
        formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M")
        self._file_name = 'Precond_'+ self._sam_name + '_' +formatted_datetime

        if (self._use_video):
            print(f"Use video. Point1: {self._point1}")
            self.signal_start_stop_tracking.emit(True)
            
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M")
            img_addr = self._workfolder + '\\Test_'+self._sam_name+'_'+formatted_datetime+'_first_frame.jpg'
            self.signal_save_image.emit(img_addr)


       

        self._start_time = time.perf_counter()
        self._start_cycle_time = time.perf_counter()
        self._current_time = 0

        #Set number of cycles and direction
        self._half_cycle = 2*self._num_cycles_precond #double for each half cycle
        self._direction = Direction.STRETCH

        #Set final positions
        self._start_force1 = 0.0
        self._start_force2 = 0.0
        self._end_force1 = self._max_force1
        self._end_force2 = self._max_force2

        self._mot_daq.zeroPosition()
        
        self._force1,self._force2 = self._mot_daq.get_forces()
        
        self._vel_ax1 = self._disp_guess1/self._test_duration
        self._vel_ax2 = self._disp_guess2/self._test_duration

        # Start slow motion
        self._mot_daq.move_velocity_ax1(-self._vel_ax1) #in mm/s
        self._mot_daq.move_velocity_ax2(-self._vel_ax2) #in mm/s


        #Start timer to periodically check length and control the test
        self._test_timer = QTimer()
        self._test_timer.timeout.connect(self.__one_cycle_labview_alg)
        self._test_timer.start(self._sample_time)
        print("Timer started")
        self.exec()

            


    def __one_cycle_labview_alg(self):

        #Condition to finish the test: positive number of steps left
        if self._current_time < 1000 and self._execute and self._half_cycle > 0:

            # Read current forces, positions, time and record them
            self._current_time = round(time.perf_counter() - self._start_time, 5)
            self._current_cycle_time = round(time.perf_counter() - self._start_cycle_time, 5)
            self._force1,self._force2 = self._mot_daq.get_forces() #try/except is inside
            #self._av_force1,self._av_force2 = self._mot_daq.get_av_forces(n = 5) #try/except is inside
            if len(self._load1) > 5:
                self._av_force1 = sum(self._load1[-5:])/5
            else:
                self._av_force1 = self._force1

            if len(self._load2) > 5:
                self._av_force2 = sum(self._load2[-5:])/5
            else:
                self._av_force2 = self._force2


            self._pos1, self._pos2 = self._mot_daq.get_positions()
            self._update_arrays_emit_data()

            #If Stretch or Relax half cycle
            if Direction.STRETCH == self._direction and self._av_force1 < self._end_force1 and self._av_force2 < self._end_force2 or \
                Direction.COMPRESS == self._direction and (self._pos1 < -0.08 or self._pos2 < -0.08):
                #(self._force1 > self._end_force1 or self._force2 > self._end_force2):
                
                
                #Stop at new 0 after very first cycle
                if self._state == State.PRECONDITIONING and self._half_cycle == 2*self._num_cycles_precond-1:
                    #If only one of the motor reached zero load - stop it
                    if Direction.COMPRESS == self._direction and self._av_force1 <= 0.1:
                        self._mot_daq.stop_motor1()
                        self._mot_daq.zero_pos1()
                       
                    if Direction.COMPRESS == self._direction and self._av_force2 <= 0.1:
                        self._mot_daq.stop_motor2()
                        self._mot_daq.zero_pos2()
                    

            #Half cycle finished
            else:
                self._half_cycle -= 1
                
                if Direction.STRETCH == self._direction:
                    self._direction = Direction.COMPRESS
                else:
                    self._direction = Direction.STRETCH
                
                #Change behaviour from relax to stretch loop
                if Direction.STRETCH == self._direction:
                    #Increase force loop
                    #print("LoadControlTest: Start increasing force")

                    self._start_force1 = 0
                    self._start_force2 = 0
                    self._end_force1 = self._max_force1
                    self._end_force2 = self._max_force2

                    self._vel_ax1 = round(self._disp_guess1/self._test_duration, 5)
                    self._vel_ax2 = round(self._disp_guess2/self._test_duration, 5)

                    self._mot_daq.stop_motors()
                    QThread.msleep(500)

                    # Start stretching
                    self._mot_daq.move_velocity_ax1(-self._vel_ax1) #in mm/s
                    self._mot_daq.move_velocity_ax2(-self._vel_ax2) #in mm/s

                    if self._state == State.PRECONDITIONING:
                        self._writeDataToFile(self._file_name, self._num_cycles_precond - self._half_cycle/2)
                    elif self._state == State.TEST:
                        self._writeDataToFile(self._file_name, self._num_cycles_test - self._half_cycle/2 )

                    self._init_variables()
                    self._start_cycle_time = time.perf_counter()

                    

                #Stretch loop finished. Change behaviour to relax loop
                else:
                    #Decrease force loop
                    #print("LoadControlTest: Start decreasing force")

                    self._start_force1 = self._max_force1
                    self._start_force2 = self._max_force2
                    self._end_force1 = 0
                    self._end_force2 = 0

                    if self._use_video and 1 == self._half_cycle:
                        #Get current date for filename
                        current_datetime = datetime.now()
                        formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M")
                        img_addr = self._workfolder + '\\Test_'+self._sam_name+'_'+formatted_datetime+'_last_frame.jpg'
                        self.signal_save_image.emit(img_addr)


                    #linear fit on the last 25% of the cycle
                    #correction calculatoin

                    n = int(0.25*self._current_cycle_time*1000/self._sample_time) #calculation of the length of the 25% of the cycle

                    #y = [x - self._start_time + self._start_cycle_time for x in self._time[-n:]] #Last 25% of the cycle
                    y = np.array(self._time_rel[-n:])
                    x1 = self._load1[-n:]
                    x2 = self._load2[-n:]

                    coef1 = np.polyfit(x1, y, 1)
                    coef2 = np.polyfit(x2, y, 1)

                    time_to_load1 = round(coef1[0]*self._max_force1 + coef1[1], 3)
                    time_to_load2 = round(coef2[0]*self._max_force2 + coef2[1], 3)

                    #corr_fact1 = round(((time_to_load1 - current_cycle_time)*0.8 + current_cycle_time) / current_cycle_time, 3)
                    #corr_fact2 = round(((time_to_load2 - current_cycle_time)*0.8 + current_cycle_time) / current_cycle_time, 3)

                    corr_fact1 = round(((time_to_load1 - self._test_duration)*0.8 + self._test_duration) / self._test_duration, 3)
                    corr_fact2 = round(((time_to_load2 - self._test_duration)*0.8 + self._test_duration) / self._test_duration, 3)

                    #print(f"time_to_load1: {time_to_load1}, time_to_load2: {time_to_load2}, current_cycle_time: {current_cycle_time}")
                    #print(f"Corr1: {corr_fact1}, Corr2: {corr_fact2}, Velocity1: {self._vel_ax1}, Velocity2: {self._vel_ax2}")
                    #print(f"Point1: {self._point1}")

                
                    corr_fact1 = min(corr_fact1, 1.5)
                    corr_fact2 = min(corr_fact2, 1.5)

                    self._disp_guess1 = self._disp_guess1 * corr_fact1
                    self._disp_guess2 = self._disp_guess2 * corr_fact2
                    

                    # Start motors
                    #self._mot_daq.move_velocity_ax1(self._vel_ax1) #in mm/s
                    #self._mot_daq.move_velocity_ax2(self._vel_ax2) #in mm/s
                    self._mot_daq.move_position_ax1(0, self._vel_ax1)
                    self._mot_daq.move_position_ax2(0, self._vel_ax2)
                

        #If cycles are over on PRECONDITIONING
        #Change state to TEST
        elif self._half_cycle <= 0 and self._state == State.PRECONDITIONING:
            self._state = State.TEST
            self._half_cycle = 2*self._num_cycles_test
            self._direction = Direction.STRETCH
            self._init_variables()
            #self._mot_daq.zeroPosition()

            print("LoadControlTest: Preconditioning finished. Main test started.")
            self.signal_precond_finished.emit()

            #Get current date for filename
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M")
            self._file_name = 'Test_'+ self._sam_name + '_' + formatted_datetime

            #Write old data to file
            #Initialize variables from the beginning
            #Initialize markers
            #Send signal to main GUI to clear the old data
            #Give a warning "Continue with a test?"
            #Start the test
        
        
        #When test finished
        else:
            # Stop motors after measurement cycle is finished
            QMetaObject.invokeMethod(self._test_timer, "stop", Qt.ConnectionType.QueuedConnection)
            self.stop_measurement()
            print("LoadControlTest: Test finished")
            self.signal_test_finished.emit()
            

         
            

        
            
            
    
